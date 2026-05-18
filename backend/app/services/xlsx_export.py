"""Build XLSX workbooks in-memory for download endpoints.

Used by the audit-log and reports exports. Keep this module tiny — anything
beyond "headers + rows → bytes" belongs in the caller.
"""

from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class Sheet:
    name: str
    headers: list[str]
    rows: Iterable[list[Any]] = field(default_factory=list)


_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="FF2D5016")  # brand dark green


def _sanitize_sheet_name(name: str) -> str:
    # Excel forbids: : \ / ? * [ ] and >31 chars
    cleaned = "".join(c for c in name if c not in r":\/?*[]")
    return cleaned[:31] or "Sheet"


def build_workbook(sheets: list[Sheet]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in sheets:
        ws = wb.create_sheet(title=_sanitize_sheet_name(sheet.name))

        if sheet.headers:
            ws.append(sheet.headers)
            for col_idx in range(1, len(sheet.headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL

        col_widths = [len(h) for h in sheet.headers]
        for row in sheet.rows:
            ws.append(row)
            for i, value in enumerate(row):
                length = len(str(value)) if value is not None else 0
                if i < len(col_widths):
                    col_widths[i] = max(col_widths[i], length)
                else:
                    col_widths.append(length)

        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 10), 60)

        if sheet.headers:
            ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
